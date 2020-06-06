
from openpyxl import load_workbook
from config import channel 
import sys
print (sys.argv[3]) # prints var2


workbook = load_workbook(filename=sys.argv[3])



sheet = workbook.active
print(len(sheet[int(sys.argv[1])]))
list_keys = []
for i in sheet[int(sys.argv[1])]:
	list_keys.append(channel.get(i.value))
	if channel.get(i.value) == None :
		b = 30 - len(i.value)
		c = ""
		x = 1
		while x < b:
			c = c +" "
			x = x + 1
		print(i.value,c,channel.get(i.value))


sheet.insert_rows(int(sys.argv[2]),1)
r = 1
cont = int(sys.argv[2])
for x in list_keys:
    sheet.cell(row=cont, column=r).value = x
    r += 1
workbook.save('C:/Users/dell/Desktop/Nouveau dossier/FileName.xlsx')